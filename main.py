import traceback
from datetime import datetime, timedelta
from asyncio import sleep, create_task
from io import BytesIO

from requests.exceptions import SSLError
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from fastapi import FastAPI, UploadFile, Request, Form
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles

from name_converter import NameExcelToDict
from card_data_conveter import CardDataConverter
from slack_communicator import Slack
from constants import WS_NEW_HEADERS

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

thread = None

@app.get("/upload-form", response_class=HTMLResponse)
async def upload_form(request: Request):
    global thread
    if thread is None or not thread.is_alive():
        templates = Jinja2Templates(directory="templates")
        return templates.TemplateResponse("upload_form.html", {"request": request})
    
    return "작업 중."

@app.post("/run/")
async def run(people_file: UploadFile, card_file: UploadFile, start_at:str, end_at:str, channel_id: str = Form() ):
    global thread
    start_at = datetime.strptime(start_at, "%Y%m%d") - timedelta(hours=9)
    end_at = datetime.strptime(end_at, "%Y%m%d") - timedelta(hours=9)
    if thread is None or not thread.is_alive():
        create_task(_run_in_thread(people_file, card_file, start_at, end_at, channel_id))
        return "작업이 시작되었습니다."
    else : 
        return "작업 중. 슬랙 메시지를 확인하세요."

async def _run_in_thread(people_file: UploadFile, card_file: UploadFile, start_at:datetime, end_at:datetime, channel_id: str = Form()):
    slack = Slack(start_at, end_at)
    send_success = slack.send_message(f"요청을 받았습니다. 시작시간 :{datetime.now()} / 검색범위 : {start_at.strftime('%Y%m%d')} ~ {end_at.strftime('%Y%m%d')}", channel_id)
    if not send_success:
        return "슬랙 메시지 전송에 실패했습니다. 채널ID를 확인해주세요."

    try:
        people_file_read = await people_file.read()
        wb_people = load_workbook(filename=BytesIO(people_file_read))
        nick_to_human, card_num_to_human = NameExcelToDict(wb_people=wb_people).run()

        # cell_list.sort()
        # cell_list = [x[1] for x in cell_list]
        # WS_NEW_HEADERS.extend(cell_list)
        WS_NEW_HEADERS.append("댓글/파일")
    # BaseException 대신 파싱하다가 날 수 있는 에러 목록 찾아서 추가 필요
    except BaseException as e:
        error_traceback = traceback.format_exc()
        slack.send_message("인명부 파일을 포맷을 확인해주세요.")
        return {"message": "인명부 파일을 포맷을 확인해주세요."}

    try:
        card_file_read = await card_file.read()
        wb_card = load_workbook(filename=BytesIO(card_file_read))
        card_data_converter = CardDataConverter(wb_card)
        card_sheet = card_data_converter.get_new_sheet()
        card_dict = card_data_converter.get_card_data_dict()
        card_data_converter.add_card_owner_to_new_sheet(card_num_to_human)
    # BaseException 대신 파싱하다가 날 수 있는 에러 목록 찾아서 추가 필요
    except BaseException as e:
        error_traceback = traceback.format_exc()
        slack.send_message("카드사 파일 포맷을 확인해주세요.")
        return {"message": "카드사 파일 포맷을 확인해주세요."}

    pay_messages = slack.crawl_all_messages()

    if not pay_messages:
        slack.error_report(message="채널에서 메시지 수집에 실패했습니다.")
        return "채널에서 메시지 수집에 실패했습니다."

    attempts = 0
    attempts += 1
    idx_passed = 0

    while attempts < 50:
        attempts += 1
        try:
            for idx, (ts, pay_channel_data) in enumerate(pay_messages.items()):
                if (idx < idx_passed) :
                    continue
                if idx % 100 == 0:
                    slack.send_message(
                        f"{len(pay_messages)}개의 글 중 {idx}번째 글의 댓글 수집중", channel_id
                    )
                if (pay_channel_data.dict_key not in card_dict) or ts is None:
                    continue
                slack.get_a_reply_from_slack(ts, pay_messages)
                row = card_dict[pay_channel_data.dict_key]

                if pay_channel_data.replies:
                    card_data_converter.add_comments(card_sheet, row, pay_channel_data)
                    card_data_converter.add_usage_and_owners(
                        card_sheet, row, pay_channel_data, nick_to_human
                    )
                idx_passed = idx
                await sleep(2)
            break
        except SSLError as ssl_error:
            await sleep(30)
            slack.error_report(message="SSL 에러", error_log=ssl_error)
        except BaseException as e:
            error_traceback = traceback.format_exc()
            slack.error_report(error_log=error_traceback)
    # 결과물 (OUTPUT)
    vb = save_virtual_workbook(wb_card)
    slack.send_file(file=vb, channel_id=channel_id)
