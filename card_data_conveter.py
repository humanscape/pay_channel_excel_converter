from string import ascii_uppercase
from openpyxl import Workbook
from constants import WS_NEW_HEADERS

from comment_converter import comment_converter


class CardDataConverter:
    def __init__(self, wb: Workbook):
        self.wb = wb
        self.ws_origin = wb["신용_국내"]
        self.row_max = self.ws_origin.max_row
        self.ROW_MIN = 2
        self.WS_EXTRA_COL = len(WS_NEW_HEADERS)
        self.USAGE_COL = WS_NEW_HEADERS.index("유형")
        self.OWNER_COL = WS_NEW_HEADERS.index("소지자")

    def get_new_sheet(self):
        self._set_additional_sheet()
        return self.ws_new

    def get_card_data_dict(self):
        return self._get_dict_time_card_amount_to_row()

    def _set_additional_sheet(self) -> Workbook.worksheets:
        self.ws_new = self._init_new_sheet()
        self._add_amount_calc_col()

    def _init_new_sheet(self):
        ws_new = self.wb.create_sheet("result")
        for idx, header in enumerate(WS_NEW_HEADERS):
            if idx == 0:
                continue
            ws_new.cell(1, idx, header)

        col_origin = 1
        while header := self.ws_origin.cell(1, col_origin).value:
            if header not in WS_NEW_HEADERS:
                col_origin += 1
                continue

            col_new = WS_NEW_HEADERS.index(header)

            for row in range(self.ROW_MIN, self.row_max + 1):
                origin_value = self.ws_origin.cell(row, col_origin).value
                ws_new.cell(row, col_new, origin_value)
            col_origin += 1

        return ws_new

    def _add_amount_calc_col(self):
        ws_new = self.ws_new
        origin_amount_col = WS_NEW_HEADERS.index("공급가액")
        full_amount_col = WS_NEW_HEADERS.index("승인금액")
        tax_col = WS_NEW_HEADERS.index("세액")
        origin_alp = ascii_uppercase[origin_amount_col - 1]
        full_col = ascii_uppercase[full_amount_col - 1]

        for row in range(self.ROW_MIN, self.row_max + 1):
            ws_new.cell(row, origin_amount_col, value=f"=ROUND({full_col}{row}/1.1,0)")
            ws_new.cell(row, tax_col, value=f"={full_col}{row}-{origin_alp}{row}")

        return ws_new

    def _get_dict_time_card_to_row(self):
        ws_new = self.ws_new
        # 시간 + 카드번호 + 금액이 겹치는 경우는 없을 것이라고 가정
        # {"2022.07.01 10:12:00 1234 15000 " : "row_num" }
        card_data_row_to_pay_data = {}
        date_col = WS_NEW_HEADERS.index("승인일")
        time_col = WS_NEW_HEADERS.index("승인시간")  # 초 00으로 변경
        card_num_col = WS_NEW_HEADERS.index("카드번호")  # 뒤 4자리만
        # amount_col = WS_NEW_HEADERS.index("승인금액")
        for row in range(self.ROW_MIN, self.row_max):
            date = ws_new.cell(row, date_col).value
            time = ws_new.cell(row, time_col).value[:-2] + "00"
            card_num = ws_new.cell(row, card_num_col).value[-4:]
            # amount = ws_new.cell(row, amount_col).value

            card_data_row_to_pay_data[f"{date} {time} {card_num}"] = row # f"{date} {time} {card_num} {amount}"

        return card_data_row_to_pay_data

    def add_pay_data_to_row(self, ws, row_num: int, data):
        return

    def add_card_owner_to_new_sheet(self, card_num_to_human):
        CIC_COL = WS_NEW_HEADERS.index("CIC")
        CELL_COL = WS_NEW_HEADERS.index("셀")
        CARD_NUM_COL = WS_NEW_HEADERS.index("카드번호")
        for row in range(self.ROW_MIN, self.row_max + 1):
            card_full_num = self.ws_new.cell(row, CARD_NUM_COL).value
            if card_full_num not in card_num_to_human:
                continue
            human = card_num_to_human[card_full_num]
            self.ws_new.cell(row, self.OWNER_COL, human.kor_name)
            self.ws_new.cell(row, CIC_COL, human.cic_name)
            self.ws_new.cell(row, CELL_COL, human.cell_name)
        return

    def add_comments(self, ws, row: int, pay_channel_data):
        for i, reply in enumerate(pay_channel_data.replies):
            ws.cell(row, self.WS_EXTRA_COL + i, reply)
        return

    def add_usage_and_owners(self, ws, row: int, pay_channel_data, nick_to_human):
        usage_and_kor_names_and_cells = comment_converter(
            pay_channel_data.replies[0], nick_to_human
        )
        usage = usage_and_kor_names_and_cells["usage"]
        kor_names = usage_and_kor_names_and_cells["kor_names"]

        if usage is not None:
            ws.cell(row, self.USAGE_COL, usage)

        if kor_names:
            ws.cell(row, self.OWNER_COL, kor_names)

        # cell_counter = usage_and_kor_names_and_cells["cell_counter"]
        # if cell_counter:
        #     for cell, count in cell_counter.items():
        #         ws.cell(row, WS_NEW_HEADERS.index(cell), int(count))

        return
