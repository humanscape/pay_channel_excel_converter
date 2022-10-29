from typing import Tuple
import openpyxl
from pydantic import BaseModel

from constants import PEOPLE_LIST_HEADERS


class Human(BaseModel):
    cic_name: str
    cell_name: str
    eng_name: str | None
    kor_name: str
    card_full_num: str | None

    def __repr__(self):
        return f"{self.cic_name} {self.cell_name} {self.eng_name}"


class NameExcelToDict:
    header_cols = {}
    nickname_cols = []
    HEADER_START_ROW = 2
    # cell_list = set()

    def __init__(self, wb_people: openpyxl.Workbook):
        self.ws = wb_people["Member List"]
        self.nick_to_human = {}
        self.card_num_to_human = {}

    def run(self) -> Tuple[dict, dict, set]:
        self._init_cols()
        self._init_dicts()

        return self.nick_to_human, self.card_num_to_human # , list(self.cell_list)

    def _init_cols(self):
        header_cells = self.ws[self.HEADER_START_ROW]
        for cell in header_cells:
            if cell.value is None:
                continue
            value = cell.value.lower()
            col_idx = cell.col_idx
            self._set_header_cols(col_idx, value)
            self._set_nickname_cols(col_idx, value)

    def _set_header_cols(self, col_idx: int, value: str) -> None:
        if value in PEOPLE_LIST_HEADERS:
            self.header_cols[PEOPLE_LIST_HEADERS[value]] = col_idx

    def _set_nickname_cols(self, col_idx: int, value: str) -> None:
        if "name" in value:
            self.nickname_cols.append(col_idx)

    def _id_exist(self, row) -> bool:
        return self.ws.cell(row, 1).value is not None

    def _init_dicts(self) -> None:
        row = self.HEADER_START_ROW + 1
        while self._id_exist(row):
            human: Human = self._set_human(row)
            self._set_nick_to_human(row, human)
            self._set_card_num_to_human(human)
            # self._set_cell_list(row)
            row += 1
        return

    def _set_human(self, row: int) -> Human:
        human_attrs = {}
        for header, col_nums in self.header_cols.items():
            human_attrs[header] = self.ws.cell(row, col_nums).value

        return Human(**human_attrs)

    def _set_nick_to_human(self, row: int, human: Human) -> None:
        for nickname_cols in self.nickname_cols:
            nickname = self.ws.cell(row, nickname_cols).value
            if nickname is None:
                continue
            if nickname in self.nick_to_human:
                self.nick_to_human[nickname] = False
                continue
            self.nick_to_human[nickname] = human
        return

    def _set_card_num_to_human(self, human: Human) -> None:
        self.card_num_to_human[human.card_full_num] = human
        return

    # def _set_cell_list(self, row:int) -> None:
    #     cell_col = self.header_cols["cell_name"]
    #     cic_col = self.header_cols["cic_name"]
    #     cell = self.ws.cell(row, cell_col).value
    #     cic = self.ws.cell(row,cic_col).value

    #     self.cell_list.add((cic,cell))
